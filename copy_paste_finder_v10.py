"""
copy_paste_finder.py  v10
-------------------------
Detects duplicated, copied and fabricated data in Excel datasets.

Runs nine strategies:
  A. Duplicate rows          – pairs of rows sharing ≥2 high-entropy values
  B. Repeated column seqs    – same run of values at two positions in a column
  C. Terminal digit test     – per-column χ²(9) uniformity test on last significant digit
  D. Periodic row duplication – fixed-lag block copying (e.g. every 101 rows)
  E. Cosine similarity       – near-identical rows on repetitive columns only
  F. Fingerprint gap         – dominant gap between recurring row fingerprints
  G. Collinearity matrix     – column pairs with |r| ≥ 0.98 (explains E false positives)
  H. Modular block count     – counts exact-match row pairs per candidate period;
                               writes output.pdf with heatmaps and autocorrelogram
  I. Excel format forensics  – file-layer analysis independent of data values:
                               internal metadata (creator/modifier/timestamps/revision),
                               Track Changes log, per-column cell font/fill/number-format
                               anomalies (Data Colada method), integer-ratio excess,
                               and column-identity pairs (one column copied from another).
                               Pages appended to --out PDF when --forensics flag is set.

Usage:
    pip install openpyxl matplotlib numpy          # core
    pip install scipy                              # optional: exact χ² p-values for Strategy C

    # Heuristic mode (strategies A–G):
    python copy_paste_finder.py <file.xlsx>

    # With Strategy H visualisation (writes output.pdf):
    python copy_paste_finder.py <file.xlsx> --plot
    python copy_paste_finder.py <file.xlsx> --plot --plot-cols WBC,Hb,Plt,BUN,Cr,Na

    # Strategy I forensic pages appended to the same PDF:
    python copy_paste_finder.py <file.xlsx> --forensics
    python copy_paste_finder.py <file.xlsx> --plot --forensics

    # All options:
    python copy_paste_finder.py <file.xlsx> [--sheet SHEET]
                                            [--min-suspicion low|medium|high]
                                            [--plot]
                                            [--plot-cols COL1,COL2,...]
                                            [--plot-period 101]
                                            [--min-period 50]
                                            [--max-period 250]
                                            [--max-lag 300]
                                            [--out output.pdf]
                                            [--forensics]
"""

from __future__ import annotations

import argparse
import math
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass, field
from enum import Enum
from typing import Optional

import openpyxl


# matplotlib is optional – only imported when --plot or --forensics is requested
_MPL_AVAILABLE: Optional[bool] = None


def _require_matplotlib() -> None:
    global _MPL_AVAILABLE
    if _MPL_AVAILABLE is None:
        try:
            import matplotlib  # noqa: F401
            _MPL_AVAILABLE = True
        except ImportError:
            _MPL_AVAILABLE = False
    if not _MPL_AVAILABLE:
        print("matplotlib and numpy are required for --plot / --forensics. "
              "Install with: pip install matplotlib numpy", file=sys.stderr)
        sys.exit(1)


# ---------------------------------------------------------------------------
# Suspicion level
# ---------------------------------------------------------------------------

class SuspicionLevel(Enum):
    NONE   = 0
    LOW    = 1
    MEDIUM = 2
    HIGH   = 3

    def __ge__(self, other): return self.value >= other.value
    def __gt__(self, other): return self.value >  other.value
    def __le__(self, other): return self.value <= other.value
    def __lt__(self, other): return self.value <  other.value


# ---------------------------------------------------------------------------
# Column profile replaces AI categorisation
# ---------------------------------------------------------------------------

@dataclass
class ColumnProfile:
    index: int
    name: str
    is_repeating_fraction: bool = False
    is_square_root: bool = False
    is_ln_argument: bool = False
    is_included: bool = True          # False for ID / grouping columns


# ---------------------------------------------------------------------------
# Fraction / ln detector
# ---------------------------------------------------------------------------

def _num_decimals(value: float) -> int:
    s = str(value)
    if "." in s:
        return len(s.rstrip("0").split(".")[1])
    return 0


def detect_repeating_fraction(value: float, tolerance: float = 1e-4) -> Optional[int]:
    """Return the numerator if value looks like p/q for small denominators, else None."""
    if not math.isfinite(value):
        return None
    if _num_decimals(value) < 4:
        return None
    denominators = [3, 7, 9, 11, 13, 17, 19, 21, 23]
    for denom in denominators:
        for k in range(4):
            tol = min(tolerance * 10 ** (-k * 1.5), 10 ** (-_num_decimals(value) + 2))
            numerator = abs(value) * denom
            rounded = round(numerator, k)
            if abs(numerator - rounded) < tol and rounded != 0:
                return round(rounded * 10 ** k)
    return None


def detect_square_root(value: float) -> Optional[float]:
    """Return radicand if value ≈ √(simple number), else None."""
    if _num_decimals(value) < 4 or value < 0.1:
        return None
    radicand = value ** 2
    rounded = round(radicand, 2)
    if abs(radicand - rounded) < 1e-5:
        return rounded
    return None


def detect_ln_argument(value: float) -> bool:
    """Heuristic: value is between 0 and 1 exclusive with many decimals → might be ln-space."""
    return 0 < value < 1 and _num_decimals(value) >= 4


# ---------------------------------------------------------------------------
# Entropy calculation
# ---------------------------------------------------------------------------

def _raw_number_entropy(value: float) -> int:
    """Strip decimal point and trailing zeros, return abs integer."""
    s = str(value).replace(".", "")
    s = s.rstrip("0") or "0"
    try:
        return abs(int(s))
    except ValueError:
        return 0


def base_number_entropy(value: float) -> int:
    """
    Compute the raw entropy of a single numeric value before column-type capping.
    """
    # Years get a fixed entropy of 100
    if 1900 <= value <= 2030 and value == int(value):
        return 100

    raw = _raw_number_entropy(value)

    # If value looks like a square root, use the radicand's entropy if lower
    radicand = detect_square_root(value)
    if radicand is not None:
        radicand_raw = _raw_number_entropy(radicand)
        if radicand_raw < raw:
            raw = radicand_raw

    # If value looks like a repeating fraction, numerator IS the entropy
    numerator = detect_repeating_fraction(value)
    if numerator is not None:
        return numerator

    # Terminating decimal check (denominators 2, 4, 8)
    candidates = [_raw_number_entropy(value * d) for d in (2, 4, 8)]
    min_candidate = min(candidates)
    if min_candidate < raw / 2:
        return min_candidate

    return raw


def number_entropy(value: float, col: ColumnProfile) -> int:
    """Apply column-type cap (matches calculateNumberEntropy)."""
    raw = base_number_entropy(value)
    if col.is_ln_argument or col.is_square_root or col.is_repeating_fraction:
        return min(raw, 100)
    return raw


def entropy_score(raw_entropy: int) -> float:
    """
    Map raw entropy to a score.
    Matches calculateEntropyScore() in entropy.ts.
    """
    if raw_entropy <= 1:
        return 0.0
    if raw_entropy < 100:
        return math.log10(raw_entropy)
    if raw_entropy < 100_000:
        return 5 * math.log10(raw_entropy) - 8
    return math.log10(raw_entropy) + 12


def column_sequence_entropy_score(values: list[float], col: ColumnProfile) -> float:
    return sum(entropy_score(number_entropy(v, col)) for v in values)


def row_entropy_score(values: list[float], cols: list[ColumnProfile]) -> float:
    return sum(entropy_score(number_entropy(v, c)) for v, c in zip(values, cols))


# ---------------------------------------------------------------------------
# Sequence regularity
# ---------------------------------------------------------------------------

def sequence_regularity(values: list[float]) -> float:
    """Return fraction of consecutive intervals that equal the most common interval."""
    if len(values) < 2:
        return 0.0
    if all(v == values[0] for v in values):
        return (len(values) - 1) / len(values)
    intervals: dict[float, int] = defaultdict(int)
    for i in range(len(values) - 1):
        intervals[values[i + 1] - values[i]] += 1
    most_common_count = max(intervals.values())
    return (most_common_count - 1) / (len(values) - 1)


# ---------------------------------------------------------------------------
# Excel reading
# ---------------------------------------------------------------------------

def _is_numeric(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool) and math.isfinite(v)


@dataclass
class SheetData:
    name: str
    rows: list[list]          # raw cell values, row-major
    headers: list[str]        # column names (row 0)
    numeric_col_indices: list[int]
    log_count_modifier: float   # normaliser = log2(total numeric cells + 2)


def read_sheet(ws) -> SheetData:
    data = list(ws.iter_rows(values_only=True))
    if not data:
        raise ValueError("Empty sheet")

    headers = [str(c) if c is not None else "" for c in data[0]]
    rows = [list(row) for row in data]

    numeric_col_indices = []
    total_numeric = 0
    for col_idx in range(len(headers)):
        for row in rows[1:]:
            if col_idx < len(row) and _is_numeric(row[col_idx]):
                total_numeric += 1
        # Mark column as numeric if ≥30% of data rows are numeric
        numeric_count = sum(
            1 for row in rows[1:]
            if col_idx < len(row) and _is_numeric(row[col_idx])
        )
        if len(rows) > 1 and numeric_count / (len(rows) - 1) >= 0.3:
            numeric_col_indices.append(col_idx)

    log_modifier = math.log2(max(total_numeric, 2))
    return SheetData(
        name=ws.title,
        rows=rows,
        headers=headers,
        numeric_col_indices=numeric_col_indices,
        log_count_modifier=log_modifier,
    )


# ---------------------------------------------------------------------------
# Column profiling (heuristic, no AI)
# ---------------------------------------------------------------------------

SAMPLE_SIZE = 10
ID_PATTERNS = re.compile(
    r"(^|_|\s)(id|identifier|no|number|code|treatment|group|"
    r"species|sex|site|plot|block|tank|cage|animal)($|_|\s)",
    re.I
)


def build_column_profiles(sheet: SheetData) -> list[ColumnProfile]:
    profiles = []
    data_rows = sheet.rows[1: 1 + SAMPLE_SIZE]

    for idx, name in enumerate(sheet.headers):
        if idx not in sheet.numeric_col_indices:
            profiles.append(ColumnProfile(index=idx, name=name, is_included=False))
            continue

        is_id = bool(ID_PATTERNS.search(name))

        sample = [r[idx] for r in data_rows if idx < len(r) and _is_numeric(r[idx])]

        is_frac = any(detect_repeating_fraction(v) is not None for v in sample)
        is_sqrt = any(detect_square_root(v) is not None for v in sample)
        is_ln   = any(detect_ln_argument(v) for v in sample)

        profiles.append(ColumnProfile(
            index=idx,
            name=name,
            is_repeating_fraction=is_frac,
            is_square_root=is_sqrt,
            is_ln_argument=is_ln,
            is_included=not is_id,
        ))
    return profiles


# ---------------------------------------------------------------------------
# Strategy A – Duplicate rows
# ---------------------------------------------------------------------------

MIN_ROW_ENTROPY_SCORE  = 50    # minimum entropy of a single shared value to seed candidates
MIN_ADJUSTED_ROW_SCORE = 4     # minimum sheet-size-adjusted entropy to report
MIN_SHARED_COLS        = 2     # minimum shared columns per pair
MAX_OCCURRENCES        = 200   # skip values that appear in too many rows (likely categorical)
MAX_DUPLICATE_ROWS     = 1000


@dataclass
class DuplicateRowResult:
    row_a: int
    row_b: int
    shared_values: list[float]
    shared_col_names: list[str]
    entropy_score: float
    adjusted_score: float
    suspicion: SuspicionLevel


def find_duplicate_rows(sheet: SheetData, profiles: list[ColumnProfile]) -> list[DuplicateRowResult]:
    included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
    if not included:
        return []

    col_indices = [p.index for p in included]
    col_by_index = {p.index: p for p in included}

    rows_by_value: dict[tuple, set[int]] = defaultdict(set)
    for row_idx in range(1, len(sheet.rows)):
        row = sheet.rows[row_idx]
        for col_idx in col_indices:
            if col_idx >= len(row):
                continue
            v = row[col_idx]
            if not _is_numeric(v):
                continue
            col = col_by_index[col_idx]
            if number_entropy(float(v), col) >= MIN_ROW_ENTROPY_SCORE:
                rows_by_value[(col_idx, v)].add(row_idx)

    compared: set[tuple[int, int]] = set()
    results: list[DuplicateRowResult] = []

    for (col_idx, _val), row_set in rows_by_value.items():
        if len(row_set) < 2 or len(row_set) > MAX_OCCURRENCES:
            continue
        row_list = sorted(row_set)
        for i in range(len(row_list)):
            for j in range(i + 1, len(row_list)):
                if len(results) >= MAX_DUPLICATE_ROWS:
                    break
                ra, rb = row_list[i], row_list[j]
                if (ra, rb) in compared:
                    continue
                compared.add((ra, rb))

                row_a = sheet.rows[ra]
                row_b = sheet.rows[rb]
                shared_vals: list[float] = []
                shared_cols: list[ColumnProfile] = []
                for ci in col_indices:
                    if ci >= len(row_a) or ci >= len(row_b):
                        continue
                    va, vb = row_a[ci], row_b[ci]
                    if _is_numeric(va) and _is_numeric(vb) and va == vb:
                        shared_vals.append(float(va))
                        shared_cols.append(col_by_index[ci])

                if len(shared_vals) < MIN_SHARED_COLS:
                    continue

                seen: set[float] = set()
                dedup_vals, dedup_cols = [], []
                for v, c in zip(shared_vals, shared_cols):
                    if v not in seen:
                        seen.add(v)
                        dedup_vals.append(v)
                        dedup_cols.append(c)

                row_ent = row_entropy_score(dedup_vals, dedup_cols)
                adjusted = row_ent / sheet.log_count_modifier

                if adjusted <= MIN_ADJUSTED_ROW_SCORE:
                    continue

                if adjusted > 16:
                    susp = SuspicionLevel.HIGH
                elif adjusted > 9:
                    susp = SuspicionLevel.MEDIUM
                elif adjusted > 6:
                    susp = SuspicionLevel.LOW
                else:
                    susp = SuspicionLevel.NONE

                results.append(DuplicateRowResult(
                    row_a=ra,
                    row_b=rb,
                    shared_values=shared_vals,
                    shared_col_names=[c.name for c in shared_cols],
                    entropy_score=row_ent,
                    adjusted_score=adjusted,
                    suspicion=susp,
                ))

    results.sort(key=lambda r: r.entropy_score, reverse=True)
    return results


# ---------------------------------------------------------------------------
# Strategy B – Repeated column sequences
# ---------------------------------------------------------------------------

MIN_SEQ_LENGTH        = 2
MIN_SEQ_ENTROPY_SCORE = 10
MIN_ADJ_SEQ_SCORE     = 2
MAX_SEQUENCES         = 1000


@dataclass
class ColumnSequenceResult:
    col_name: str
    col_index: int
    row_a: int        # 0-based row index (including header)
    row_b: int
    values: list[float]
    seq_entropy: float
    adjusted_entropy: float
    matrix_adjusted: float
    suspicion: SuspicionLevel


def find_repeated_sequences(sheet: SheetData, profiles: list[ColumnProfile]) -> list[ColumnSequenceResult]:
    included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
    if not included:
        return []

    results: list[ColumnSequenceResult] = []
    checked_pairs: set[tuple] = set()

    for col_prof in included:
        ci = col_prof.index
        col_vals = [
            (ri, float(sheet.rows[ri][ci]))
            if ci < len(sheet.rows[ri]) and _is_numeric(sheet.rows[ri][ci])
            else (ri, None)
            for ri in range(len(sheet.rows))
        ]

        positions_by_value: dict[float, list[int]] = defaultdict(list)
        for ri, v in col_vals:
            if v is not None:
                positions_by_value[v].append(ri)

        for start_v, row_indices in positions_by_value.items():
            if len(row_indices) < 2 or len(row_indices) > MAX_OCCURRENCES:
                continue
            for i in range(len(row_indices)):
                for j in range(i + 1, len(row_indices)):
                    if len(results) >= MAX_SEQUENCES:
                        break
                    ra, rb = row_indices[i], row_indices[j]
                    if ra == rb:
                        continue
                    pair_key = (ci, ra, rb)
                    if pair_key in checked_pairs:
                        continue

                    run: list[float] = [start_v]
                    length = 1
                    while True:
                        next_a = ra + length
                        next_b = rb + length
                        if next_a >= len(col_vals) or next_b >= len(col_vals):
                            break
                        va = col_vals[next_a][1]
                        vb = col_vals[next_b][1]
                        if va is None or vb is None or va != vb:
                            break
                        if ra == ci and rb == ra + length:
                            break
                        run.append(va)
                        checked_pairs.add((ci, ra + length, rb + length))
                        length += 1

                    if len(run) < MIN_SEQ_LENGTH:
                        continue

                    seq_ent = column_sequence_entropy_score(run, col_prof)
                    if seq_ent <= MIN_SEQ_ENTROPY_SCORE:
                        continue

                    regularity = sequence_regularity(run)
                    adj_ent = seq_ent * (1 - regularity)
                    matrix_adj = adj_ent / sheet.log_count_modifier

                    if matrix_adj <= MIN_ADJ_SEQ_SCORE:
                        continue

                    if matrix_adj > 16:
                        susp = SuspicionLevel.HIGH
                    elif matrix_adj > 9:
                        susp = SuspicionLevel.MEDIUM
                    elif matrix_adj > 7:
                        susp = SuspicionLevel.LOW
                    else:
                        susp = SuspicionLevel.NONE

                    results.append(ColumnSequenceResult(
                        col_name=col_prof.name,
                        col_index=ci,
                        row_a=ra,
                        row_b=rb,
                        values=run,
                        seq_entropy=seq_ent,
                        adjusted_entropy=adj_ent,
                        matrix_adjusted=matrix_adj,
                        suspicion=susp,
                    ))
                    checked_pairs.add(pair_key)

    results.sort(key=lambda r: r.matrix_adjusted, reverse=True)
    seen_scores: set[tuple] = set()
    deduped = []
    for r in results:
        key = (r.col_index, round(r.adjusted_entropy, 6), tuple(r.values))
        if key not in seen_scores:
            seen_scores.add(key)
            deduped.append(r)
    return deduped


# ---------------------------------------------------------------------------
# Strategy C – Terminal digit frequency test (per-column χ²)
# ---------------------------------------------------------------------------
# For each included numeric column, extract the last significant (non-trailing-
# zero) digit of every value and run a χ²(9) goodness-of-fit test against the
# uniform distribution over {0, …, 9}.  Fabricated data often shows non-uniform
# terminal digits (humans avoid 7 and 9, over-use 0, 2, 5).
#
# Columns with repeating-fraction values are skipped (their terminal digits are
# mathematically constrained and would produce spurious hits).
# Requires ≥ C_MIN_VALUES qualifying values per column.
# scipy is used for the p-value if available; otherwise a Wilson-Hilferty normal
# approximation is used.

C_MIN_VALUES = 30    # minimum qualifying values per column
C_P_HIGH     = 0.001
C_P_MEDIUM   = 0.01
C_P_LOW      = 0.05


def _terminal_digit(value: float) -> Optional[int]:
    """
    Return the last significant (non-trailing-zero) decimal digit of *value*,
    or None when the value is an integer or has no fractional part.
    """
    s = str(value)
    if "." not in s:
        return None
    frac = s.split(".", 1)[1].rstrip("0")
    if not frac:
        return None          # whole number stored as float
    return int(frac[-1])


def _chi2_uniform_p(counts: list[int]) -> tuple[float, float]:
    """
    χ²(9) goodness-of-fit against U{0,…,9}.
    Returns (chi2_stat, p_value).
    """
    n = sum(counts)
    if n == 0:
        return 0.0, 1.0
    expected = n / 10.0
    chi2_stat = sum((c - expected) ** 2 / expected for c in counts)
    try:
        from scipy.stats import chi2 as scipy_chi2
        p = float(scipy_chi2.sf(chi2_stat, df=9))
    except (ImportError, ValueError):
        # Wilson-Hilferty normal approximation for χ²(df=9)
        df = 9
        z = ((chi2_stat / df) ** (1 / 3) - (1 - 2 / (9 * df))) / math.sqrt(2 / (9 * df))
        p = 0.5 * math.erfc(z / math.sqrt(2))
    return chi2_stat, p


@dataclass
class TerminalDigitResult:
    col_name:     str
    col_index:    int
    digit_counts: list[int]   # counts for digits 0–9
    chi2:         float
    p_value:      float
    n_values:     int
    suspicion:    SuspicionLevel


def find_terminal_digit_anomalies(
    sheet: SheetData,
    profiles: list[ColumnProfile],
) -> list[TerminalDigitResult]:
    # Skip repeating-fraction columns – their terminal digits are constrained
    included = [
        p for p in profiles
        if p.is_included
        and p.index in sheet.numeric_col_indices
        and not p.is_repeating_fraction
    ]
    if not included:
        return []

    data_rows = sheet.rows[1:]
    results: list[TerminalDigitResult] = []

    for col_prof in included:
        ci = col_prof.index
        counts = [0] * 10
        n_valid = 0

        for row in data_rows:
            if ci >= len(row):
                continue
            v = row[ci]
            if not _is_numeric(v):
                continue
            fv = float(v)
            # skip year-like integers
            if 1900 <= fv <= 2030 and fv == int(fv):
                continue
            d = _terminal_digit(fv)
            if d is None:
                continue
            counts[d] += 1
            n_valid += 1

        if n_valid < C_MIN_VALUES:
            continue

        chi2_stat, p = _chi2_uniform_p(counts)

        if p < C_P_HIGH:
            susp = SuspicionLevel.HIGH
        elif p < C_P_MEDIUM:
            susp = SuspicionLevel.MEDIUM
        elif p < C_P_LOW:
            susp = SuspicionLevel.LOW
        else:
            continue    # NONE – don't record

        results.append(TerminalDigitResult(
            col_name=col_prof.name,
            col_index=ci,
            digit_counts=counts,
            chi2=round(chi2_stat, 3),
            p_value=round(p, 6),
            n_values=n_valid,
            suspicion=susp,
        ))

    results.sort(key=lambda r: r.chi2, reverse=True)
    return results


# ---------------------------------------------------------------------------
# Strategy D – Periodic row duplication (fixed-lag block copying)
# ---------------------------------------------------------------------------
# Detects the BMJ pattern: blocks of rows copied with a fixed period p,
# so row i and row i+p share many column values across the whole dataset.

MIN_PERIOD       = 10    # smallest lag to test
MAX_PERIOD       = 500   # largest lag to test
MIN_COLS_MATCH   = 5     # shared numeric columns required per row pair
MIN_PAIR_HITS    = 3     # how many (i, i+p) pairs must match to flag a period


@dataclass
class PeriodicDuplicationResult:
    period: int
    matching_pairs: int      # number of (row_i, row_i+period) pairs that match
    example_rows: list[tuple[int, int]]   # up to 5 example pairs
    matched_col_names: list[str]          # columns that most often match
    suspicion: SuspicionLevel


def find_periodic_duplications(
    sheet: SheetData,
    profiles: list[ColumnProfile],
) -> list[PeriodicDuplicationResult]:
    included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
    if not included:
        return []

    col_indices = [p.index for p in included]
    data_rows = sheet.rows[1:]   # skip header; index 0 here = sheet row 1
    n = len(data_rows)

    results: list[PeriodicDuplicationResult] = []

    for period in range(MIN_PERIOD, min(MAX_PERIOD + 1, n)):
        pairs_matched = 0
        example_pairs: list[tuple[int, int]] = []
        col_hit_counts: dict[int, int] = defaultdict(int)

        for i in range(n - period):
            row_a = data_rows[i]
            row_b = data_rows[i + period]
            shared = 0
            for ci in col_indices:
                if ci >= len(row_a) or ci >= len(row_b):
                    continue
                va, vb = row_a[ci], row_b[ci]
                if _is_numeric(va) and _is_numeric(vb) and va == vb:
                    shared += 1
                    col_hit_counts[ci] += 1
            if shared >= MIN_COLS_MATCH:
                pairs_matched += 1
                if len(example_pairs) < 5:
                    example_pairs.append((i + 1, i + 1 + period))  # 1-based sheet rows

        if pairs_matched < MIN_PAIR_HITS:
            continue

        # suspicion scales with how many pairs matched relative to what's possible
        ratio = pairs_matched / (n - period)
        if ratio > 0.15 or pairs_matched >= 10:
            susp = SuspicionLevel.HIGH
        elif ratio > 0.05 or pairs_matched >= 5:
            susp = SuspicionLevel.MEDIUM
        else:
            susp = SuspicionLevel.LOW

        top_cols = sorted(col_hit_counts, key=lambda c: col_hit_counts[c], reverse=True)[:8]
        matched_col_names = [sheet.headers[c] for c in top_cols if c < len(sheet.headers)]

        results.append(PeriodicDuplicationResult(
            period=period,
            matching_pairs=pairs_matched,
            example_rows=example_pairs,
            matched_col_names=matched_col_names,
            suspicion=susp,
        ))

    results.sort(key=lambda r: r.matching_pairs, reverse=True)
    return results


# ---------------------------------------------------------------------------
# Strategy E – Cosine similarity on high-repetition columns only
# ---------------------------------------------------------------------------
# Uses only columns where the same value appears in ≥ REPEAT_THRESHOLD fraction
# of rows (i.e. columns with excess repeated values).
# This prevents non-copied columns (IDs, dates) from diluting similarity scores.

COSINE_THRESHOLD      = 0.999   # similarity above this is flagged
MAX_ROWS_FOR_COSINE   = 2000    # skip full O(n²) pairwise above this
MIN_SIMILAR_PAIRS     = 3
REPEAT_COL_THRESHOLD  = 0.02    # column must have ≥ this fraction of repeated values


def _repetitive_col_indices(sheet: SheetData, profiles: list[ColumnProfile]) -> list[int]:
    """Return indices of included columns that have excess repeated values."""
    included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
    data_rows = sheet.rows[1:]
    n = len(data_rows)
    if n == 0:
        return []

    repetitive = []
    for p in included:
        ci = p.index
        counts: dict[float, int] = defaultdict(int)
        for row in data_rows:
            v = row[ci] if ci < len(row) else None
            if _is_numeric(v):
                counts[float(v)] += 1
        # fraction of values that appear more than once
        repeated = sum(c for c in counts.values() if c > 1)
        if repeated / n >= REPEAT_COL_THRESHOLD:
            repetitive.append(ci)
    return repetitive


@dataclass
class CosineSimilarityResult:
    similar_pairs: list[tuple[int, int, float]]   # (row_a, row_b, similarity), 1-based
    num_pairs: int
    cols_used: list[str]
    suspicion: SuspicionLevel


def find_cosine_similar_rows(
    sheet: SheetData,
    profiles: list[ColumnProfile],
) -> Optional[CosineSimilarityResult]:
    col_indices = _repetitive_col_indices(sheet, profiles)
    if len(col_indices) < 3:
        return None

    data_rows = sheet.rows[1:]
    n = len(data_rows)

    if n > MAX_ROWS_FOR_COSINE:
        print(f"  [Strategy E] {n} rows exceeds limit {MAX_ROWS_FOR_COSINE}, skipping")
        return None

    vectors: list[list[float]] = []
    for row in data_rows:
        vec = [float(row[ci]) if ci < len(row) and _is_numeric(row[ci]) else 0.0
               for ci in col_indices]
        vectors.append(vec)

    def cosine(a: list[float], b: list[float]) -> float:
        dot = sum(x * y for x, y in zip(a, b))
        na  = math.sqrt(sum(x * x for x in a))
        nb  = math.sqrt(sum(x * x for x in b))
        return dot / (na * nb) if na > 0 and nb > 0 else 0.0

    similar_pairs: list[tuple[int, int, float]] = []
    for i in range(n):
        for j in range(i + 1, n):
            sim = cosine(vectors[i], vectors[j])
            if sim >= COSINE_THRESHOLD:
                similar_pairs.append((i + 1, j + 1, round(sim, 6)))
            if len(similar_pairs) > 5000:
                break
        if len(similar_pairs) > 5000:
            break

    if len(similar_pairs) < MIN_SIMILAR_PAIRS:
        return None

    count = len(similar_pairs)
    susp = SuspicionLevel.HIGH if count >= 20 else SuspicionLevel.MEDIUM if count >= 8 else SuspicionLevel.LOW
    cols_used = [sheet.headers[ci] for ci in col_indices if ci < len(sheet.headers)]

    return CosineSimilarityResult(
        similar_pairs=similar_pairs[:50],
        num_pairs=count,
        cols_used=cols_used,
        suspicion=susp,
    )


# ---------------------------------------------------------------------------
# Strategy F – Fingerprint-gap on included measurement columns only
# ---------------------------------------------------------------------------
# Builds a row fingerprint from Claude-screened (or heuristic-screened) included
# columns only, excluding IDs and grouping variables. A dominant gap between
# recurring fingerprints reveals the block-copy period.

MIN_FINGERPRINT_RECURRENCES = 2
MIN_DOMINANT_GAP_FRACTION   = 0.3   # gap must account for ≥ this share of all gaps
MIN_COL_COMPLETENESS        = 0.7   # column must be ≥ this fraction non-null to fingerprint


@dataclass
class FingerprintGapResult:
    dominant_gap: int
    gap_count: int
    total_gaps: int
    gap_fraction: float
    example_row_pairs: list[tuple[int, int]]
    cols_used: list[str]
    suspicion: SuspicionLevel


def _is_monotonic(sheet: SheetData, col_idx: int) -> bool:
    """Return True if a column's values are strictly monotonically increasing."""
    vals = [sheet.rows[i][col_idx] for i in range(1, len(sheet.rows))
            if col_idx < len(sheet.rows[i]) and _is_numeric(sheet.rows[i][col_idx])]
    if len(vals) < 4:
        return False
    return all(vals[i] < vals[i + 1] for i in range(len(vals) - 1))


def find_fingerprint_gaps(
    sheet: SheetData,
    profiles: list[ColumnProfile],
) -> Optional[FingerprintGapResult]:
    included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
    if not included:
        return None

    data_rows = sheet.rows[1:]
    n = len(data_rows)
    if n == 0:
        return None

    # Detect obs_id: search ALL numeric columns (including excluded ones like 'obs_id')
    # for a unique integer column used as the observation ordering variable.
    # Gaps must be computed in obs_id space because data may not be sorted by template.
    all_numeric = [p for p in profiles if p.index in sheet.numeric_col_indices]
    obs_id_col: Optional[int] = None
    for p in all_numeric:
        ci = p.index
        vals = [float(row[ci]) for row in data_rows
                if ci < len(row) and _is_numeric(row[ci])]
        if not vals:
            continue
        if (all(v == int(v) for v in vals) and
                len(set(vals)) / len(vals) > 0.95):
            obs_id_col = ci
            break

    obs_ids = [
        int(row[obs_id_col]) if obs_id_col is not None and obs_id_col < len(row)
        and _is_numeric(row[obs_id_col]) else (i + 1)
        for i, row in enumerate(data_rows)
    ]

    # Keep complete, non-monotonic, low-cardinality columns for fingerprinting.
    # unique_frac < 0.25 excludes outcome variables and high-precision measurements
    # that differ between copied rows even when lab values are identical.
    col_indices = []
    for p in included:
        ci = p.index
        if ci == obs_id_col:
            continue
        if _is_monotonic(sheet, ci):
            continue
        vals = [float(row[ci]) for row in data_rows
                if ci < len(row) and _is_numeric(row[ci])]
        if not vals or len(vals) / n < MIN_COL_COMPLETENESS:
            continue
        if len(set(vals)) / len(vals) < 0.25:
            col_indices.append(ci)

    if len(col_indices) < MIN_COLS_MATCH:
        # Not enough low-cardinality columns found automatically.
        # Use --ai mode for Claude column screening, or pass --cols explicitly
        # to plot_duplicates.py for the modular block count visualization.
        return None

    # Build fingerprint → list of obs_id values
    fingerprint_obs: dict[tuple, list[int]] = defaultdict(list)
    for i, row in enumerate(data_rows):
        fp = tuple(
            float(row[ci]) if ci < len(row) and _is_numeric(row[ci]) else None
            for ci in col_indices
        )
        if sum(v is not None for v in fp) < MIN_COLS_MATCH:
            continue
        fingerprint_obs[fp].append(obs_ids[i])

    all_gaps: list[int] = []
    gap_pairs: list[tuple[int, int]] = []
    for obs_id_list in fingerprint_obs.values():
        if len(obs_id_list) < MIN_FINGERPRINT_RECURRENCES:
            continue
        sorted_ids = sorted(obs_id_list)
        for k in range(len(sorted_ids) - 1):
            gap = sorted_ids[k + 1] - sorted_ids[k]
            all_gaps.append(gap)
            gap_pairs.append((sorted_ids[k], sorted_ids[k + 1]))

    if not all_gaps:
        return None

    gap_counts: dict[int, int] = defaultdict(int)
    for g in all_gaps:
        gap_counts[g] += 1

    dominant_gap, dominant_count = max(gap_counts.items(), key=lambda x: x[1])
    total_gaps = len(all_gaps)
    fraction = dominant_count / total_gaps

    if fraction < MIN_DOMINANT_GAP_FRACTION:
        return None

    susp = (SuspicionLevel.HIGH   if fraction > 0.7 and dominant_count >= 5 else
            SuspicionLevel.MEDIUM if fraction > 0.4 or  dominant_count >= 3 else
            SuspicionLevel.LOW)

    examples = [(a, b) for a, b in gap_pairs if b - a == dominant_gap][:5]
    cols_used = [sheet.headers[ci] for ci in col_indices if ci < len(sheet.headers)]

    return FingerprintGapResult(
        dominant_gap=dominant_gap,
        gap_count=dominant_count,
        total_gaps=total_gaps,
        gap_fraction=round(fraction, 3),
        example_row_pairs=examples,
        cols_used=cols_used,
        suspicion=susp,
    )


# ---------------------------------------------------------------------------
# Strategy G – Collinearity matrix
# ---------------------------------------------------------------------------
# Computes Pearson r between every pair of included numeric columns.
# Column pairs with |r| ≥ COLLINEARITY_THRESHOLD are reported as collinear.
# A cluster of mutually collinear columns (e.g. L*, a*, b*, X, Y, Z, C*, h)
# explains high Strategy E similarity scores and marks them as likely false
# positives rather than copied data.

COLLINEARITY_THRESHOLD   = 0.98   # |r| above this = collinear
MIN_COLLINEAR_PAIRS      = 1      # need at least this many pairs to report


@dataclass
class CollinearPair:
    col_a: str
    col_b: str
    r: float


@dataclass
class CollinearityResult:
    pairs: list[CollinearPair]       # all collinear pairs, sorted by |r| desc
    collinear_col_names: list[str]   # deduplicated column names involved
    independent_col_names: list[str] # included columns NOT in any collinear pair


def _pearson_r(xs: list[float], ys: list[float]) -> Optional[float]:
    n = len(xs)
    if n < 4:
        return None
    mx = sum(xs) / n
    my = sum(ys) / n
    num = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    dx  = math.sqrt(sum((x - mx) ** 2 for x in xs))
    dy  = math.sqrt(sum((y - my) ** 2 for y in ys))
    if dx == 0 or dy == 0:
        return None
    return num / (dx * dy)


def find_collinear_columns(
    sheet: SheetData,
    profiles: list[ColumnProfile],
) -> Optional[CollinearityResult]:
    included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
    if len(included) < 2:
        return None

    data_rows = sheet.rows[1:]

    # Extract column vectors
    col_vecs: dict[int, list[float]] = {}
    for p in included:
        ci = p.index
        vec = [float(row[ci]) for row in data_rows
               if ci < len(row) and _is_numeric(row[ci])]
        if len(vec) >= 4:
            col_vecs[ci] = vec

    pairs: list[CollinearPair] = []
    indices = list(col_vecs.keys())
    for i in range(len(indices)):
        for j in range(i + 1, len(indices)):
            ci, cj = indices[i], indices[j]
            xs, ys = col_vecs[ci], col_vecs[cj]
            # align lengths (missing values may differ)
            length = min(len(xs), len(ys))
            r = _pearson_r(xs[:length], ys[:length])
            if r is not None and abs(r) >= COLLINEARITY_THRESHOLD:
                pairs.append(CollinearPair(
                    col_a=sheet.headers[ci] if ci < len(sheet.headers) else str(ci),
                    col_b=sheet.headers[cj] if cj < len(sheet.headers) else str(cj),
                    r=round(r, 4),
                ))

    if len(pairs) < MIN_COLLINEAR_PAIRS:
        return None

    pairs.sort(key=lambda p: abs(p.r), reverse=True)

    collinear_names: list[str] = []
    seen: set[str] = set()
    for p in pairs:
        for name in (p.col_a, p.col_b):
            if name not in seen:
                seen.add(name)
                collinear_names.append(name)

    all_included_names = {
        sheet.headers[p.index] for p in included if p.index < len(sheet.headers)
    }
    independent_names = sorted(all_included_names - seen)

    return CollinearityResult(
        pairs=pairs,
        collinear_col_names=collinear_names,
        independent_col_names=independent_names,
    )


# ---------------------------------------------------------------------------
# Strategy H – Modular block count + visualisation (output.pdf)
# ---------------------------------------------------------------------------
# For each candidate period p, counts how many row pairs at obs_id distance p
# share identical values in ≥ MIN_COLS_MATCH columns (NaN-ignoring).
# The true copy period produces a sharp peak. Data need not be sorted by obs_id.
# Also writes a PDF with four pages:
#   1. Z-score heatmap sorted by obs_id
#   2. Z-score heatmap sorted by obs_id mod <period>
#   3. Modular block count bar chart
#   4. Exact-match autocorrelogram on obs_id-sorted data

H_MIN_PERIOD    = 50
H_MAX_PERIOD    = 250
H_MAX_LAG       = 300
H_MIN_MATCH_COLS = 5


@dataclass
class ModularBlockResult:
    period: int
    block_count: int
    counts_by_period: list[tuple[int, int]]   # (period, count) for all candidates
    suspicion: SuspicionLevel


def _h_obs_id_array(sheet: SheetData) -> list[int]:
    """Return obs_id values (all rows, 1-based if no obs_id col found)."""
    data_rows = sheet.rows[1:]
    for ci in sheet.numeric_col_indices:
        vals = [sheet.rows[i][ci] for i in range(1, len(sheet.rows))
                if ci < len(sheet.rows[i]) and _is_numeric(sheet.rows[i][ci])]
        if vals and all(v == int(v) for v in vals) and len(set(vals)) / len(vals) > 0.95:
            return [int(sheet.rows[i][ci]) if ci < len(sheet.rows[i]) and _is_numeric(sheet.rows[i][ci])
                    else i for i in range(1, len(sheet.rows))]
    return list(range(1, len(data_rows) + 1))


def find_modular_blocks(
    sheet: SheetData,
    profiles: list[ColumnProfile],
    min_period: int = H_MIN_PERIOD,
    max_period: int = H_MAX_PERIOD,
    plot_cols: Optional[list[str]] = None,
) -> Optional[ModularBlockResult]:
    """
    Compute modular block count for candidate periods.
    Uses plot_cols if supplied, otherwise falls back to AI-included columns,
    then to all complete low-cardinality numeric columns.
    """
    import numpy as np

    data_rows = sheet.rows[1:]
    n = len(data_rows)
    if n < min_period * 2:
        return None

    # Determine which columns to use
    if plot_cols:
        header_to_idx = {h: i for i, h in enumerate(sheet.headers)}
        col_indices = [header_to_idx[c] for c in plot_cols if c in header_to_idx]
    else:
        included = [p for p in profiles if p.is_included and p.index in sheet.numeric_col_indices]
        col_indices = [p.index for p in included
                       if not _is_monotonic(sheet, p.index)
                       and sum(1 for row in data_rows
                               if p.index < len(row) and _is_numeric(row[p.index])) / n >= MIN_COL_COMPLETENESS
                       and len(set(float(row[p.index]) for row in data_rows
                                   if p.index < len(row) and _is_numeric(row[p.index]))) / n < 0.25]

    if len(col_indices) < H_MIN_MATCH_COLS:
        return None

    # Build obs_id-sorted matrix
    obs_ids = _h_obs_id_array(sheet)
    order = sorted(range(n), key=lambda i: obs_ids[i])
    matrix = np.array([
        [float(data_rows[i][ci]) if ci < len(data_rows[i]) and _is_numeric(data_rows[i][ci])
         else float("nan") for ci in col_indices]
        for i in order
    ], dtype=float)

    # Modular block count
    max_p = min(max_period, n // 2)
    counts_arr = np.zeros(max_p - min_period + 1, dtype=int)
    for pi, p in enumerate(range(min_period, max_p + 1)):
        for i in range(n - p):
            a, b = matrix[i], matrix[i + p]
            mask = ~(np.isnan(a) | np.isnan(b))
            if mask.sum() >= H_MIN_MATCH_COLS and np.all(a[mask] == b[mask]):
                counts_arr[pi] += 1

    if counts_arr.max() == 0:
        return None

    best_period = int(np.argmax(counts_arr)) + min_period
    best_count  = int(counts_arr[best_period - min_period])
    counts_by_period = [(min_period + i, int(c)) for i, c in enumerate(counts_arr)]

    n_possible = n - best_period
    ratio = best_count / n_possible if n_possible > 0 else 0
    susp = (SuspicionLevel.HIGH   if ratio > 0.3 or best_count >= 20 else
            SuspicionLevel.MEDIUM if ratio > 0.1 or best_count >= 5  else
            SuspicionLevel.LOW)

    return ModularBlockResult(
        period=best_period,
        block_count=best_count,
        counts_by_period=counts_by_period,
        suspicion=susp,
    )


def run_strategy_h_plot(
    sheet: SheetData,
    profiles: list[ColumnProfile],
    result: Optional[ModularBlockResult],
    pdf,
    plot_cols: Optional[list[str]],
    min_period: int,
    max_period: int,
    max_lag: int,
) -> None:
    """Append Strategy H diagnostic pages to an open PdfPages object."""
    import numpy as np
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors

    CMAP = mcolors.TwoSlopeNorm(vmin=-3, vcenter=0, vmax=3)
    data_rows = sheet.rows[1:]
    n = len(data_rows)

    if n == 0:
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.text(0.5, 0.5, "Empty sheet: no data rows to plot",
                ha="center", va="center", transform=ax.transAxes, fontsize=10)
        ax.set_axis_off()
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)
        return

    # Determine display columns (all continuous non-binary)
    if plot_cols:
        header_to_idx = {h: i for i, h in enumerate(sheet.headers)}
        disp_indices = [header_to_idx[c] for c in plot_cols if c in header_to_idx]
        disp_names   = [c for c in plot_cols if c in header_to_idx]
    else:
        disp_indices, disp_names = [], []
        for ci in sheet.numeric_col_indices:
            vals = [float(row[ci]) for row in data_rows
                    if ci < len(row) and _is_numeric(row[ci])]
            if vals and not all(v in (0.0, 1.0) for v in vals):
                disp_indices.append(ci)
                disp_names.append(sheet.headers[ci] if ci < len(sheet.headers) else str(ci))

    obs_ids = _h_obs_id_array(sheet)
    order = sorted(range(n), key=lambda i: obs_ids[i])
    obs_sorted = np.array([obs_ids[i] for i in order], dtype=float)

    period = result.period if result else None

    def info_page(title, message):
        fig, ax = plt.subplots(figsize=(10, 4))
        ax.text(0.5, 0.5, message, ha="center", va="center",
                transform=ax.transAxes, fontsize=10)
        ax.set_title(title, fontsize=10, pad=8)
        ax.set_axis_off()
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def heatmap(ax, z_data, sort_col, title, ylabels):
        if sort_col is not None:
            z_data = z_data[np.argsort(sort_col)]
        im = ax.imshow(np.where(np.isnan(z_data), 0.0, z_data).T, aspect="auto",
                       cmap="RdBu_r", norm=CMAP, interpolation="nearest")
        ax.set_yticks(range(len(ylabels)))
        ax.set_yticklabels(ylabels, fontsize=7)
        ax.set_xticks([])
        ax.set_title(title, fontsize=10, pad=4)
        plt.colorbar(im, ax=ax, fraction=0.02, pad=0.01, label="Z-score")

    # Build display matrix only if display columns exist
    if disp_indices:
        raw = np.array([
            [float(data_rows[i][ci]) if ci < len(data_rows[i]) and _is_numeric(data_rows[i][ci])
             else float("nan") for ci in disp_indices]
            for i in order
        ], dtype=float).reshape(len(order), len(disp_indices))

        z = np.full_like(raw, np.nan)
        for ci in range(raw.shape[1]):
            col = raw[:, ci]
            valid = col[~np.isnan(col)]
            if len(valid) > 1:
                mu, sd = valid.mean(), valid.std()
                if sd > 0:
                    z[:, ci] = (col - mu) / sd

        fig, ax = plt.subplots(figsize=(10, max(4, len(disp_names) * 0.22 + 1)))
        heatmap(ax, z, None, "Full dataset – sorted by obs_id", disp_names)
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

        if period:
            fig, ax = plt.subplots(figsize=(10, max(4, len(disp_names) * 0.22 + 1)))
            heatmap(ax, z, obs_sorted % period,
                    f"Sorted by obs_id mod {period} – copies should align", disp_names)
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)
    else:
        info_page("Strategy H – heatmap",
                  "No eligible continuous non-binary numeric columns to display.")
        if period:
            info_page("Strategy H – mod-period heatmap",
                      f"Period {period} detected, but no eligible display columns were available.")

    # Page 3: modular block count
    fig, ax = plt.subplots(figsize=(12, 4))
    if result and result.counts_by_period:
        periods_x = [p for p, _ in result.counts_by_period]
        counts_y  = [c for _, c in result.counts_by_period]
        ax.bar(periods_x, counts_y, width=1.0, color="#5B9BD5", linewidth=0, alpha=0.8)
        if period:
            ax.axvline(period, color="#C0392B", linewidth=1.5, linestyle="--",
                       label=f"period={period}  n={result.block_count}")
        if counts_y:
            top = int(np.argmax(counts_y))
            ax.annotate(f"p={periods_x[top]}\nn={counts_y[top]}",
                        xy=(periods_x[top], counts_y[top]),
                        xytext=(periods_x[top] + (max(periods_x) - min(periods_x)) * 0.02,
                                counts_y[top]),
                        fontsize=8, color="#C0392B",
                        arrowprops=dict(arrowstyle="->", color="#C0392B", lw=0.8))
    else:
        ax.text(0.5, 0.5, "No block count computed\n(too few qualifying columns)",
                ha="center", va="center", transform=ax.transAxes, fontsize=10)
    ax.set_xlabel("Candidate period (rows)", fontsize=9)
    ax.set_ylabel("Exact-match row pairs", fontsize=9)
    ax.set_title("Strategy H – modular block count", fontsize=10, pad=4)
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(fontsize=8, frameon=False)
    fig.tight_layout()
    pdf.savefig(fig)
    plt.close(fig)

    # Autocorrelogram on complete columns
    if plot_cols:
        header_to_idx = {h: i for i, h in enumerate(sheet.headers)}
        comp_indices = [header_to_idx[c] for c in plot_cols if c in header_to_idx]
    else:
        comp_indices = [ci for ci in disp_indices
                        if sum(1 for row in data_rows if ci < len(row) and _is_numeric(row[ci])) / n >= 0.7]

    if not comp_indices or min(max_lag, n // 2) < 1:
        info_page("Strategy H – autocorrelogram",
                  "Not enough qualifying columns or rows to compute autocorrelogram.")
        return

    comp_matrix = np.array([
        [float(data_rows[i][ci]) if ci < len(data_rows[i]) and _is_numeric(data_rows[i][ci])
         else float("nan") for ci in comp_indices]
        for i in order
    ], dtype=float).reshape(len(order), len(comp_indices))

    sims = np.zeros(min(max_lag, n // 2))
    for k in range(1, len(sims) + 1):
        vals = []
        for i in range(n - k):
            a, b = comp_matrix[i], comp_matrix[i + k]
            mask = ~(np.isnan(a) | np.isnan(b))
            if mask.sum() >= 2:
                vals.append(float(np.mean(a[mask] == b[mask])))
        sims[k - 1] = float(np.mean(vals)) if vals else 0.0

    fig, ax = plt.subplots(figsize=(12, 4))
    lags = np.arange(1, len(sims) + 1)
    baseline = float(np.median(sims))
    ax.bar(lags, sims, width=1.0, color="#5B9BD5", linewidth=0, alpha=0.8)
    ax.axhline(baseline, color="#888", linewidth=0.8, linestyle="--",
               label=f"median={baseline:.4f}")
    margin = max((sims.max() - sims.min()) * 0.15, 0.002)
    ax.set_ylim(sims.min() - margin, sims.max() + margin)
    if period and period > 0:
        for mult in range(1, len(sims) // period + 1):
            lag = period * mult
            if lag <= len(sims):
                ax.axvline(lag, color="#C0392B", linewidth=1.2, linestyle="--", alpha=0.8,
                           label=f"period×{mult}={lag}" if mult <= 3 else None)
    top3 = np.argsort(sims)[-3:][::-1]
    for idx in top3:
        ax.annotate(f"lag={idx+1}\n{sims[idx]:.4f}",
                    xy=(idx + 1, sims[idx]),
                    xytext=(idx + 1 + len(sims) * 0.02, sims[idx]),
                    fontsize=7, color="#C0392B",
                    arrowprops=dict(arrowstyle="->", color="#C0392B", lw=0.8))
    ax.set_xlabel("Lag (rows, sorted by obs_id)", fontsize=9)
    ax.set_ylabel("Mean exact-match fraction", fontsize=9)
    ax.set_title("Strategy H – autocorrelogram on obs_id-sorted data", fontsize=10, pad=4)
    ax.text(0.01, 0.02, f"columns: {len(comp_indices)}",
            transform=ax.transAxes, fontsize=7, color="#888")
    handles, labels = ax.get_legend_handles_labels()
    if handles:
        ax.legend(fontsize=8, frameon=False)
    ax.set_xlim(0, len(sims) + 1)
    fig.tight_layout()
    pdf.savefig(fig)
    plt.close(fig)

# ---------------------------------------------------------------------------
# Strategy I – Excel format forensics (file-layer, not data-layer)
# ---------------------------------------------------------------------------
# Operates on the raw ZIP structure and openpyxl cell formatting objects.
# Detects signals invisible to strategies A–H, which work only on cell values:
#
#   I-1  Internal metadata   creator, last_modified_by, created/modified
#                            timestamps, revision counter (docProps/core.xml)
#   I-2  Track Changes log  presence/absence of xl/revisions/revisionLog*.xml;
#                            if present: change count and authors
#   I-3  Font anomalies     per-column: cells whose font name/size/bold/colour
#                            differs from the column-dominant signature (the core
#                            Data Colada method – copy-pasted cells retain their
#                            source font)
#   I-4  Fill anomalies     per-column: cells whose background fill differs from
#                            the column-dominant fill
#   I-5  Numfmt anomalies   per-column: mixed number formats in the same column
#   I-6  Integer ratio      per-column and global: fraction of numeric values that
#                            are exact integers; >85% in a continuous-measure column
#                            is suspicious
#   I-7  Column identity    pairs of columns where ≥ I_IDENTITY_THRESHOLD of rows
#                            have identical values – one was likely copied from the
#                            other (the #B==$B signal in the Gino tax study)
#
# Requires: openpyxl (already a dependency), matplotlib (already a dependency),
#           zipfile + xml.etree.ElementTree (stdlib).

import zipfile as _zipfile
import xml.etree.ElementTree as _ET
import datetime as _datetime
import hashlib as _hashlib
from collections import Counter as _Counter

I_IDENTITY_THRESHOLD = 0.95   # fraction of rows that must agree to flag identity


# ── I helpers ────────────────────────────────────────────────────────────────

def _i_read_zip_xml(path: str, member: str):
    try:
        with _zipfile.ZipFile(path) as z:
            with z.open(member) as f:
                return _ET.parse(f).getroot()
    except (KeyError, _ET.ParseError):
        return None


_I_NS = {
    "cp":      "http://schemas.openxmlformats.org/package/2006/metadata/core-properties",
    "dc":      "http://purl.org/dc/elements/1.1/",
    "dcterms": "http://purl.org/dc/terms/",
    "app":     "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties",
}


def _i_metadata(path: str) -> dict:
    meta: dict = {}
    root = _i_read_zip_xml(path, "docProps/core.xml")
    if root is not None:
        def t(tag):
            el = root.find(tag, _I_NS)
            return el.text.strip() if el is not None and el.text else None
        meta["creator"]          = t("dc:creator")
        meta["last_modified_by"] = t("cp:lastModifiedBy")
        meta["created"]          = t("dcterms:created")
        meta["modified"]         = t("dcterms:modified")
        meta["revision"]         = t("cp:revision")
    root2 = _i_read_zip_xml(path, "docProps/app.xml")
    if root2 is not None:
        def ta(tag):
            el = root2.find(tag, _I_NS)
            return el.text.strip() if el is not None and el.text else None
        meta["application"] = ta("app:Application")
        meta["app_version"] = ta("app:AppVersion")
        meta["company"]     = ta("app:Company")
    return meta


def _i_revisions(path: str) -> dict:
    result = {"revision_files": [], "change_count": 0, "authors": []}
    with _zipfile.ZipFile(path) as z:
        rev_files = [m for m in z.namelist() if m.startswith("xl/revisions/")]
        result["revision_files"] = rev_files
        ns_r = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        for rf in rev_files:
            try:
                with z.open(rf) as f:
                    rroot = _ET.parse(f).getroot()
                for rcc in rroot.iter(f"{{{ns_r}}}rcc"):
                    result["change_count"] += 1
                    author = rcc.get("author", "")
                    if author and author not in result["authors"]:
                        result["authors"].append(author)
            except Exception:
                pass
    return result


def _i_cell_font_sig(cell) -> str:
    f = cell.font
    if f is None:
        return "default"
    try:
        rgb = f.color.rgb if f.color and f.color.type == "rgb" else "none"
    except Exception:
        rgb = "none"
    return f"{f.name}|{f.size}|{f.bold}|{f.italic}|{rgb}"


def _i_cell_fill_sig(cell) -> str:
    fi = cell.fill
    if fi is None or fi.fill_type in (None, "none"):
        return "none"
    try:
        fg = fi.fgColor.rgb if fi.fgColor and fi.fgColor.type == "rgb" else "none"
        return f"{fi.fill_type}|{fg}"
    except Exception:
        return "unknown"


def _i_cell_numfmt_sig(cell) -> str:
    return str(cell.number_format) if cell.number_format else "General"


@dataclass
class IColumnFormatAnomaly:
    sheet:        str
    column:       str          # column letter (A, B, …)
    col_name:     str          # header name
    kind:         str          # "font" | "fill" | "numfmt"
    variants:     int          # number of distinct signatures
    outlier_rows: list[int]    # row numbers with non-dominant signature
    total_cells:  int
    dominant:     str


@dataclass
class IColumnIntegerRatio:
    sheet:    str
    col_name: str
    ratio:    float
    n:        int


@dataclass
class IColumnIdentityPair:
    sheet:    str
    col_a:    str
    col_b:    str
    ratio:    float      # fraction of rows where value_a == value_b
    n:        int


@dataclass
class IForensicsResult:
    path:              str
    metadata:          dict
    revisions:         dict
    format_anomalies:  list[IColumnFormatAnomaly]
    integer_ratios:    list[IColumnIntegerRatio]   # only >85% columns
    identity_pairs:    list[IColumnIdentityPair]
    flags:             list[dict]   # {"severity": HIGH/WARNING/INFO/OK, "msg": str}


def _i_year(iso: Optional[str]) -> Optional[int]:
    if not iso:
        return None
    try:
        return int(iso[:4])
    except (ValueError, TypeError):
        return None


def run_strategy_i(path: str) -> IForensicsResult:
    """
    Run all Strategy I checks on *path*.
    Loads the workbook a second time with data_only=False to access cell
    formatting objects (openpyxl loses formatting in data_only=True mode).
    """
    from openpyxl.utils import get_column_letter

    meta  = _i_metadata(path)
    revs  = _i_revisions(path)

    # Load with formatting preserved
    wb_fmt = openpyxl.load_workbook(path, data_only=False)
    wb_dat = openpyxl.load_workbook(path, data_only=True)

    format_anomalies: list[IColumnFormatAnomaly] = []
    integer_ratios:   list[IColumnIntegerRatio]   = []
    identity_pairs:   list[IColumnIdentityPair]   = []

    for sname in wb_fmt.sheetnames:
        ws_fmt = wb_fmt[sname]
        ws_dat = wb_dat[sname]

        # Per-column collections
        col_fonts:   dict[int, list] = defaultdict(list)
        col_fills:   dict[int, list] = defaultdict(list)
        col_numfmts: dict[int, list] = defaultdict(list)
        col_values:  dict[int, list] = defaultdict(list)   # for integer ratio + identity

        for row_f, row_d in zip(ws_fmt.iter_rows(), ws_dat.iter_rows()):
            for cell_f, cell_d in zip(row_f, row_d):
                if cell_f.value is None and cell_d.value is None:
                    continue
                c = cell_f.column
                r = cell_f.row
                # Formatting from the unforced load
                if cell_f.value is not None:
                    col_fonts[c].append((r, _i_cell_font_sig(cell_f)))
                    col_fills[c].append((r, _i_cell_fill_sig(cell_f)))
                    col_numfmts[c].append((r, _i_cell_numfmt_sig(cell_f)))
                # Numeric values from data_only load
                if isinstance(cell_d.value, (int, float)) and not isinstance(cell_d.value, bool):
                    if math.isfinite(cell_d.value):
                        col_values[c].append(cell_d.value)

        # Header row for column names
        headers_row = [cell.value for cell in next(ws_dat.iter_rows(max_row=1))]
        def col_header(col_idx_1based: int) -> str:
            idx = col_idx_1based - 1
            if 0 <= idx < len(headers_row):
                v = headers_row[idx]
                return str(v) if v is not None else get_column_letter(col_idx_1based)
            return get_column_letter(col_idx_1based)

        # Format anomalies (skip header row = row 1)
        for kind, col_data in (("font", col_fonts), ("fill", col_fills),
                                ("numfmt", col_numfmts)):
            for col_idx, items in col_data.items():
                data_items = [(r, s) for r, s in items if r > 1]
                if len(data_items) < 3:
                    continue
                sigs = [s for _, s in data_items]
                counts = _Counter(sigs)
                if len(counts) < 2:
                    continue
                dominant = counts.most_common(1)[0][0]
                outliers = [r for r, s in data_items if s != dominant]
                format_anomalies.append(IColumnFormatAnomaly(
                    sheet=sname,
                    column=get_column_letter(col_idx),
                    col_name=col_header(col_idx),
                    kind=kind,
                    variants=len(counts),
                    outlier_rows=outliers[:20],
                    total_cells=len(data_items),
                    dominant=dominant[:80],
                ))

        # Integer ratio per column (flag if >85% integers, n>=10)
        for col_idx, vals in col_values.items():
            if len(vals) < 10:
                continue
            n_int = sum(1 for v in vals if v == int(v))
            ratio = n_int / len(vals)
            if ratio > 0.85:
                integer_ratios.append(IColumnIntegerRatio(
                    sheet=sname,
                    col_name=col_header(col_idx),
                    ratio=round(ratio, 4),
                    n=len(vals),
                ))

        # Column identity pairs
        numeric_cols = sorted(col_values.keys())
        for i in range(len(numeric_cols)):
            for j in range(i + 1, len(numeric_cols)):
                ci, cj = numeric_cols[i], numeric_cols[j]
                vals_i, vals_j = col_values[ci], col_values[cj]
                n = min(len(vals_i), len(vals_j))
                if n < 5:
                    continue
                matches = sum(1 for a, b in zip(vals_i[:n], vals_j[:n]) if a == b)
                ratio = matches / n
                if ratio >= I_IDENTITY_THRESHOLD:
                    identity_pairs.append(IColumnIdentityPair(
                        sheet=sname,
                        col_a=col_header(ci),
                        col_b=col_header(cj),
                        ratio=round(ratio, 4),
                        n=n,
                    ))

    flags = _i_build_flags(meta, revs, format_anomalies, integer_ratios, identity_pairs)
    return IForensicsResult(
        path=path,
        metadata=meta,
        revisions=revs,
        format_anomalies=format_anomalies,
        integer_ratios=integer_ratios,
        identity_pairs=identity_pairs,
        flags=flags,
    )


def _i_build_flags(meta, revs, fmt_anomalies, int_ratios, id_pairs) -> list[dict]:
    flags = []

    # Metadata
    creator  = meta.get("creator", "") or ""
    modifier = meta.get("last_modified_by", "") or ""
    created  = meta.get("created", "") or ""
    modified = meta.get("modified", "") or ""

    if creator and modifier and creator != modifier:
        flags.append({"severity": "INFO",
                      "msg": f"Created by '{creator}' but last modified by '{modifier}'."})

    yr_c, yr_m = _i_year(created), _i_year(modified)
    if yr_c and yr_m and yr_m - yr_c >= 3:
        sev = "HIGH" if yr_m - yr_c >= 5 else "WARNING"
        flags.append({"severity": sev,
                      "msg": f"Metadata timestamp gap: created {yr_c}, modified {yr_m} "
                             f"({yr_m - yr_c} years). Late modification by '{modifier}' is suspicious."})

    rev = meta.get("revision")
    if rev and int(rev) == 1:
        flags.append({"severity": "WARNING",
                      "msg": "Revision counter = 1 — file was saved only once or metadata was reset."})

    if created and modified and created == modified:
        flags.append({"severity": "WARNING",
                      "msg": "Created and modified timestamps are identical — metadata may have been reset."})

    # Track Changes
    if not revs["revision_files"]:
        flags.append({"severity": "INFO",
                      "msg": "No Track Changes revision log found. Edit history unavailable."})
    elif revs["change_count"] > 0:
        flags.append({"severity": "INFO",
                      "msg": f"Track Changes log present: {revs['change_count']} changes, "
                             f"authors: {', '.join(revs['authors']) or 'unknown'}."})

    # Font anomalies — the Data Colada signal
    font_cols = [a for a in fmt_anomalies if a.kind == "font"]
    if font_cols:
        # Compute fraction-of-outliers for each and pick the most severe
        worst = max(font_cols, key=lambda a: a.outlier_rows.__len__())
        sev = "HIGH" if len(font_cols) >= 10 else "WARNING"
        flags.append({"severity": sev,
                      "msg": f"{len(font_cols)} column(s) with mixed font signatures — "
                             f"suggests copy-paste from different source. "
                             f"Worst: sheet '{worst.sheet}' col {worst.column} "
                             f"({len(worst.outlier_rows)} outlier rows)."})

    fill_cols = [a for a in fmt_anomalies if a.kind == "fill"]
    if fill_cols:
        flags.append({"severity": "WARNING",
                      "msg": f"{len(fill_cols)} column(s) with inconsistent cell background fills."})

    numfmt_cols = [a for a in fmt_anomalies if a.kind == "numfmt"]
    if numfmt_cols:
        flags.append({"severity": "WARNING",
                      "msg": f"{len(numfmt_cols)} column(s) with mixed number formats."})

    # Integer ratio excess
    for ir in int_ratios:
        flags.append({"severity": "WARNING",
                      "msg": f"Sheet '{ir.sheet}' col '{ir.col_name}': "
                             f"{ir.ratio*100:.0f}% of {ir.n} values are exact integers — "
                             f"suspicious for continuous measurements."})

    # Column identity pairs
    for ip in id_pairs:
        flags.append({"severity": "HIGH",
                      "msg": f"Sheet '{ip.sheet}': column '{ip.col_a}' and '{ip.col_b}' "
                             f"are identical in {ip.ratio*100:.1f}% of {ip.n} rows — "
                             f"one was likely copied from the other."})

    if not flags:
        flags.append({"severity": "OK", "msg": "No suspicious patterns detected by Strategy I."})

    return flags


def print_strategy_i(result: IForensicsResult, min_suspicion: SuspicionLevel) -> None:
    """Print Strategy I findings to stdout."""
    SEV_ORDER = {"OK": 0, "INFO": 1, "WARNING": 2, "HIGH": 3}
    threshold = {SuspicionLevel.LOW: 1, SuspicionLevel.MEDIUM: 2, SuspicionLevel.HIGH: 3}
    min_sev = threshold.get(min_suspicion, 1)

    print(f"\n[Strategy I] Excel format forensics")

    m = result.metadata
    print(f"  creator:   {m.get('creator', '—')}")
    print(f"  modifier:  {m.get('last_modified_by', '—')}")
    print(f"  created:   {m.get('created', '—')}")
    print(f"  modified:  {m.get('modified', '—')}")
    print(f"  revision:  {m.get('revision', '—')}")
    rev = result.revisions
    if rev["revision_files"]:
        print(f"  track-changes: {rev['change_count']} changes  "
              f"authors: {', '.join(rev['authors']) or 'none'}")
    else:
        print("  track-changes: none")

    filtered_flags = [f for f in result.flags if SEV_ORDER.get(f["severity"], 0) >= min_sev]
    SEV_EMOJI = {"OK": "✅", "INFO": "ℹ️", "WARNING": "⚠️", "HIGH": "🔴"}
    for f in filtered_flags:
        print(f"  {SEV_EMOJI.get(f['severity'], '?')} [{f['severity']}] {f['msg']}")

    font_count = sum(1 for a in result.format_anomalies if a.kind == "font")
    fill_count = sum(1 for a in result.format_anomalies if a.kind == "fill")
    nf_count   = sum(1 for a in result.format_anomalies if a.kind == "numfmt")
    print(f"  format anomalies: font={font_count}  fill={fill_count}  numfmt={nf_count}")
    print(f"  integer-ratio flags: {len(result.integer_ratios)}")
    print(f"  column identity pairs: {len(result.identity_pairs)}")



def write_forensics_pages(result: IForensicsResult, pdf) -> None:
    """Append Strategy I forensic pages to an open PdfPages object (matplotlib)."""
    import matplotlib.pyplot as plt
    import datetime as _dt

    SEV_COLOR = {"OK": "#27ae60", "INFO": "#2980b9", "WARNING": "#e67e22", "HIGH": "#c0392b"}

    def _text_page(title, lines, line_colors=None):
        fig, ax = plt.subplots(figsize=(12, 8))
        ax.axis("off")
        ax.set_title(title, fontsize=13, fontweight="bold", color="#2c3e50",
                     loc="left", pad=12)
        y = 0.95
        for i, line in enumerate(lines):
            col = (line_colors[i] if line_colors and i < len(line_colors) else "#222222")
            ax.text(0.01, y, line, transform=ax.transAxes, fontsize=8,
                    color=col, va="top", fontfamily="monospace")
            y -= 0.042
            if y < 0.02:
                break
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    def _table_page(title, col_headers, rows):
        fig, ax = plt.subplots(figsize=(12, max(3, min(len(rows) * 0.32 + 1.5, 9))))
        ax.axis("off")
        ax.set_title(title, fontsize=11, fontweight="bold", color="#2c3e50",
                     loc="left", pad=8)
        if not rows:
            ax.text(0.5, 0.5, "None detected.", ha="center", va="center",
                    transform=ax.transAxes, fontsize=9, color="#888")
        else:
            t = ax.table(cellText=rows[:40], colLabels=col_headers,
                         loc="center", cellLoc="left")
            t.auto_set_font_size(False)
            t.set_fontsize(7.5)
            for (r, c), cell in t.get_celld().items():
                if r == 0:
                    cell.set_facecolor("#2c3e50")
                    cell.set_text_props(color="white", fontweight="bold")
                elif r % 2 == 0:
                    cell.set_facecolor("#f5f5f5")
                cell.set_edgecolor("#cccccc")
            t.scale(1, 1.4)
        fig.tight_layout()
        pdf.savefig(fig)
        plt.close(fig)

    m   = result.metadata
    rev = result.revisions
    filename = os.path.basename(result.path)

    # ── Page: flags + metadata ──────────────────────────────────────────────
    lines, colors_ = [], []
    lines.append(f"File: {filename}   Generated: {_dt.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    colors_.append("#555555")
    lines.append("")
    colors_.append("#ffffff")
    lines.append("FORENSIC FLAGS")
    colors_.append("#2c3e50")
    lines.append("─" * 90)
    colors_.append("#cccccc")
    for f in result.flags:
        lines.append(f"  [{f['severity']:7s}]  {f['msg']}")
        colors_.append(SEV_COLOR.get(f["severity"], "#222"))
    lines += ["", "METADATA", "─" * 90]
    colors_ += ["#ffffff", "#2c3e50", "#cccccc"]
    for k, v in [
        ("Creator",          m.get("creator",          "—")),
        ("Last Modified By", m.get("last_modified_by", "—")),
        ("Created",          m.get("created",          "—")),
        ("Modified",         m.get("modified",         "—")),
        ("Revision",         m.get("revision",         "—")),
        ("Application",      m.get("application",      "—")),
        ("App Version",      m.get("app_version",      "—")),
        ("Company",          m.get("company",          "—")),
        ("Track Changes log",
         str(rev["revision_files"]) if rev["revision_files"] else "absent"),
        ("Recorded changes", str(rev["change_count"])),
        ("Change authors",   ", ".join(rev["authors"]) or "—"),
    ]:
        lines.append(f"  {k:<22s}  {v}")
        colors_.append("#222222")
    _text_page("Strategy I – Excel Format Forensics", lines, colors_)

    # ── Page: font anomalies ────────────────────────────────────────────────
    font_rows = [
        [a.sheet, a.column, a.col_name[:28], str(a.variants),
         str(a.outlier_rows[:4])[1:-1], a.dominant[:44]]
        for a in result.format_anomalies if a.kind == "font"
    ]
    _table_page("I-3  Font Anomalies – Data Colada method",
                ["Sheet", "Col", "Header", "Variants", "Outlier rows (sample)", "Dominant sig"],
                font_rows)

    # ── Page: fill + numfmt anomalies ───────────────────────────────────────
    other_rows = (
        [[a.sheet, a.column, a.col_name[:28], "fill",   str(a.variants), str(a.outlier_rows[:3])[1:-1]]
         for a in result.format_anomalies if a.kind == "fill"] +
        [[a.sheet, a.column, a.col_name[:28], "numfmt", str(a.variants), str(a.outlier_rows[:3])[1:-1]]
         for a in result.format_anomalies if a.kind == "numfmt"]
    )
    _table_page("I-4/5  Fill & Number-Format Anomalies",
                ["Sheet", "Col", "Header", "Kind", "Variants", "Outlier rows"], other_rows)

    # ── Page: integer ratios ────────────────────────────────────────────────
    ir_rows = [[ir.sheet, ir.col_name, f"{ir.ratio*100:.1f}%", str(ir.n)]
               for ir in result.integer_ratios]
    _table_page("I-6  Integer-Ratio Excess (>85% exact integers in numeric column)",
                ["Sheet", "Column", "Integer %", "n"], ir_rows)

    # ── Page: identity pairs ────────────────────────────────────────────────
    id_rows = [[ip.sheet, ip.col_a, ip.col_b, f"{ip.ratio*100:.1f}%", str(ip.n)]
               for ip in result.identity_pairs]
    _table_page("I-7  Column Identity Pairs (one column likely copied from another)",
                ["Sheet", "Column A", "Column B", "Match %", "n"], id_rows)

# ---------------------------------------------------------------------------
# Report printing
# ---------------------------------------------------------------------------

SUSPICION_EMOJI = {
    SuspicionLevel.NONE:   "⬜",
    SuspicionLevel.LOW:    "🟡",
    SuspicionLevel.MEDIUM: "🟠",
    SuspicionLevel.HIGH:   "🔴",
}


def print_report(
    sheet: SheetData,
    dup_rows: list[DuplicateRowResult],
    sequences: list[ColumnSequenceResult],
    terminal_digits: list[TerminalDigitResult],
    periodic: list[PeriodicDuplicationResult],
    cosine: Optional[CosineSimilarityResult],
    fingerprint: Optional[FingerprintGapResult],
    collinearity: Optional[CollinearityResult],
    modular: Optional[ModularBlockResult],
    min_suspicion: SuspicionLevel = SuspicionLevel.LOW,
) -> None:
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet.name!r}  |  rows: {len(sheet.rows)-1}  |  log_modifier: {sheet.log_count_modifier:.2f}")
    print(f"{'='*60}")

    # --- Strategy A ---
    filtered_rows = [r for r in dup_rows if r.suspicion >= min_suspicion]
    print(f"\n[Strategy A] Duplicate rows  ({len(filtered_rows)} findings)")
    for r in filtered_rows[:20]:
        vals_str = ", ".join(str(v) for v in r.shared_values[:6])
        if len(r.shared_values) > 6:
            vals_str += " …"
        cols_str = ", ".join(r.shared_col_names[:10])
        if len(r.shared_col_names) > 10:
            cols_str += " …"
        print(f"  {SUSPICION_EMOJI[r.suspicion]} rows {r.row_a}↔{r.row_b}  "
              f"shared_cols={len(r.shared_values)}  adj={r.adjusted_score:.2f}  "
              f"[{cols_str}]")
        print(f"     values: {vals_str}")

    # --- Strategy B ---
    filtered_seqs = [r for r in sequences if r.suspicion >= min_suspicion]
    print(f"\n[Strategy B] Repeated column sequences  ({len(filtered_seqs)} findings)")
    for r in filtered_seqs[:20]:
        vals_str = ", ".join(str(v) for v in r.values[:8])
        if len(r.values) > 8:
            vals_str += " …"
        print(f"  {SUSPICION_EMOJI[r.suspicion]} col={r.col_name!r}  "
              f"rows {r.row_a}↔{r.row_b}  len={len(r.values)}  "
              f"matrix_adj={r.matrix_adjusted:.2f}")
        print(f"     values: [{vals_str}]")

    # --- Strategy C ---
    filtered_digits = [r for r in terminal_digits if r.suspicion >= min_suspicion]
    print(f"\n[Strategy C] Terminal digit test  ({len(filtered_digits)} columns flagged)")
    for r in filtered_digits[:20]:
        bar = "".join(
            f"{d}:{r.digit_counts[d]} " for d in range(10)
        ).rstrip()
        print(f"  {SUSPICION_EMOJI[r.suspicion]} col={r.col_name!r}  "
              f"n={r.n_values}  χ²={r.chi2:.2f}  p={r.p_value:.6f}")
        print(f"     digits: {bar}")

    # --- Strategy D ---
    filtered_periodic = [r for r in periodic if r.suspicion >= min_suspicion]
    print(f"\n[Strategy D] Periodic row duplication  ({len(filtered_periodic)} findings)")
    for r in filtered_periodic[:10]:
        pairs_str = ", ".join(f"{a}↔{b}" for a, b in r.example_rows[:3])
        print(f"  {SUSPICION_EMOJI[r.suspicion]} period={r.period}  "
              f"matching_pairs={r.matching_pairs}")
        print(f"     example rows: {pairs_str}")
        print(f"     columns: {', '.join(r.matched_col_names[:6])}")

    # --- Strategy E ---
    print(f"\n[Strategy E] Cosine similarity", end="")
    if cosine is None:
        print("  (skipped or no findings)")
    elif cosine.suspicion < min_suspicion:
        print(f"  (below threshold: {cosine.num_pairs} similar pairs found)")
    else:
        print(f"  ({cosine.num_pairs} similar pairs, {len(cosine.cols_used)} repetitive columns)")
        print(f"     columns used: {', '.join(cosine.cols_used[:8])}")
        for row_a, row_b, sim in cosine.similar_pairs[:10]:
            print(f"  {SUSPICION_EMOJI[cosine.suspicion]} rows {row_a}↔{row_b}  sim={sim:.6f}")

    # --- Strategy F ---
    print(f"\n[Strategy F] Fingerprint gap", end="")
    if fingerprint is None:
        print("  (no dominant gap found)")
    elif fingerprint.suspicion < min_suspicion:
        print(f"  (below threshold: dominant gap={fingerprint.dominant_gap}, "
              f"fraction={fingerprint.gap_fraction:.2f})")
    else:
        pairs_str = ", ".join(f"{a}↔{b}" for a, b in fingerprint.example_row_pairs[:3])
        print(f"\n  {SUSPICION_EMOJI[fingerprint.suspicion]} "
              f"dominant gap={fingerprint.dominant_gap}  "
              f"occurrences={fingerprint.gap_count}/{fingerprint.total_gaps}  "
              f"fraction={fingerprint.gap_fraction:.2f}")
        print(f"     example rows: {pairs_str}")
        print(f"     columns used: {', '.join(fingerprint.cols_used[:8])}")

    # --- Strategy G ---
    print(f"\n[Strategy G] Collinearity matrix", end="")
    if collinearity is None:
        print("  (no collinear pairs found)")
    else:
        e_warning = ""
        if cosine and cosine.num_pairs >= MIN_SIMILAR_PAIRS:
            e_warning = "  ← explains Strategy E false positives"
        print(f"  ({len(collinearity.pairs)} collinear pairs ≥ |r|={COLLINEARITY_THRESHOLD}){e_warning}")
        for p in collinearity.pairs[:15]:
            print(f"     r={p.r:+.4f}  {p.col_a!r} ↔ {p.col_b!r}")
        if collinearity.independent_col_names:
            print(f"     independent columns: {', '.join(collinearity.independent_col_names[:10])}")

    # --- Strategy H ---
    print(f"\n[Strategy H] Modular block count", end="")
    if modular is None:
        print("  (not run – use --plot or too few qualifying columns)")
    elif modular.suspicion < min_suspicion:
        print(f"  (below threshold: period={modular.period}, count={modular.block_count})")
    else:
        print(f"\n  {SUSPICION_EMOJI[modular.suspicion]} "
              f"period={modular.period}  block_count={modular.block_count}")

    any_finding = (
        filtered_rows or filtered_seqs or filtered_digits or filtered_periodic
        or (cosine and cosine.suspicion >= min_suspicion)
        or (fingerprint and fingerprint.suspicion >= min_suspicion)
        or collinearity
        or (modular and modular.suspicion >= min_suspicion)
    )
    if not any_finding:
        print("\n  ✅ No suspicious findings at this threshold.")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def analyse_file(
    path: str,
    sheet_name: Optional[str] = None,
    min_suspicion: SuspicionLevel = SuspicionLevel.LOW,
    plot: bool = False,
    plot_cols: Optional[list[str]] = None,
    plot_period: Optional[int] = None,
    min_period: int = H_MIN_PERIOD,
    max_period: int = H_MAX_PERIOD,
    max_lag: int = H_MAX_LAG,
    out: str = "output.pdf",
    forensics: bool = False,
) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target_sheets = [sheet_name] if sheet_name else wb.sheetnames

    # Open a single PdfPages for the whole run when H or I (or both) need output
    pdf_obj = None
    if plot or plot_cols or forensics:
        _require_matplotlib()
        import matplotlib
        matplotlib.use("Agg")
        from matplotlib.backends.backend_pdf import PdfPages
        pdf_obj = PdfPages(out)

    try:
        for sname in target_sheets:
            if sname not in wb.sheetnames:
                print(f"Sheet {sname!r} not found. Available: {wb.sheetnames}",
                      file=sys.stderr)
                continue
            ws = wb[sname]
            try:
                sheet = read_sheet(ws)
            except ValueError as e:
                print(f"Skipping sheet {sname!r}: {e}", file=sys.stderr)
                continue

            profiles = build_column_profiles(sheet)

            dup_rows        = find_duplicate_rows(sheet, profiles)
            sequences       = find_repeated_sequences(sheet, profiles)
            terminal_digits = find_terminal_digit_anomalies(sheet, profiles)
            periodic        = find_periodic_duplications(sheet, profiles)
            cosine          = find_cosine_similar_rows(sheet, profiles)
            fingerprint     = find_fingerprint_gaps(sheet, profiles)
            collinearity    = find_collinear_columns(sheet, profiles)

            # Strategy H: modular block count + heatmap pages
            modular: Optional[ModularBlockResult] = None
            if (plot or plot_cols) and pdf_obj is not None:
                modular = find_modular_blocks(
                    sheet, profiles,
                    min_period=min_period, max_period=max_period,
                    plot_cols=plot_cols,
                )
                if modular and plot_period:
                    modular = ModularBlockResult(
                        period=plot_period,
                        block_count=dict(modular.counts_by_period).get(plot_period, 0),
                        counts_by_period=modular.counts_by_period,
                        suspicion=modular.suspicion,
                    )
                run_strategy_h_plot(
                    sheet, profiles, modular, pdf_obj, plot_cols,
                    min_period, max_period, max_lag,
                )

            print_report(sheet, dup_rows, sequences, terminal_digits,
                         periodic, cosine, fingerprint, collinearity,
                         modular, min_suspicion)

        # Strategy I file-level forensics pages once per file, after all sheets
        if forensics and pdf_obj is not None:
            try:
                i_result = run_strategy_i(path)
                print_strategy_i(i_result, min_suspicion)
                write_forensics_pages(i_result, pdf_obj)
            except KeyError as e:
                print(f"Strategy I skipped: malformed XLSX archive ({e})", file=sys.stderr)
            except Exception as e:
                print(f"Strategy I skipped due to unexpected error: {e}", file=sys.stderr)

    finally:
        if pdf_obj is not None:
            pdf_obj.close()
            print(f"  PDF written to: {out}")


def main():
    parser = argparse.ArgumentParser(description="Copy-paste finder")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--sheet", help="Sheet name (default: all sheets)")
    parser.add_argument(
        "--min-suspicion", choices=["low", "medium", "high"],
        default="low", help="Minimum suspicion level to report (default: low)"
    )
    parser.add_argument(
        "--plot", action="store_true",
        help="Run Strategy H and write output.pdf with heatmaps and block-count chart"
    )
    parser.add_argument(
        "--plot-cols", default=None,
        help="Comma-separated column names for Strategy H (e.g. WBC,Hb,Plt,BUN,Cr,Na)"
    )
    parser.add_argument(
        "--plot-period", type=int, default=None,
        help="Force a specific period for Strategy H plots (auto-detected if omitted)"
    )
    parser.add_argument(
        "--min-period", type=int, default=H_MIN_PERIOD,
        help=f"Minimum candidate period for Strategy H (default {H_MIN_PERIOD})"
    )
    parser.add_argument(
        "--max-period", type=int, default=H_MAX_PERIOD,
        help=f"Maximum candidate period for Strategy H (default {H_MAX_PERIOD})"
    )
    parser.add_argument(
        "--max-lag", type=int, default=H_MAX_LAG,
        help=f"Maximum lag for Strategy H autocorrelogram (default {H_MAX_LAG})"
    )
    parser.add_argument(
        "--out", default="output.pdf",
        help="Output PDF path for --plot / --forensics (default: output.pdf)"
    )
    parser.add_argument(
        "--forensics", action="store_true",
        help="Run Strategy I and append forensic pages to --out PDF "
             "(requires matplotlib)"
    )
    args = parser.parse_args()

    plot_cols = [c.strip() for c in args.plot_cols.split(",")] if args.plot_cols else None

    level_map = {"low": SuspicionLevel.LOW, "medium": SuspicionLevel.MEDIUM, "high": SuspicionLevel.HIGH}
    analyse_file(
        args.file,
        sheet_name=args.sheet,
        min_suspicion=level_map[args.min_suspicion],
        plot=args.plot,
        plot_cols=plot_cols,
        plot_period=args.plot_period,
        min_period=args.min_period,
        max_period=args.max_period,
        max_lag=args.max_lag,
        out=args.out,
        forensics=args.forensics,
    )


if __name__ == "__main__":
    main()
